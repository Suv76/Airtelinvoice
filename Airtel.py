import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

def process_data(file_path):
    try:
        data = pd.read_csv(file_path)
        if data.empty:
            raise ValueError("The file contains no data.")


        total_amount = data['Net Amount Payable(CR)'].sum()
        commercial_value = total_amount * 0.0025
        GST = commercial_value * 0.18
        invoice = GST + commercial_value

        summary_data = {
            'Description': ['Agent Total', '0.25%', '18%', 'Total Invoice'],
            'Payout': [total_amount, commercial_value, GST, invoice]
        }
        summary_df = pd.DataFrame(summary_data)

        file_path = 'Airtel.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='Summary')
            data.to_excel(writer, index=False, sheet_name='Data')

        wb = load_workbook(file_path)
        ws = wb['Summary']

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True, size=14)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = header_fill

        row_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if cell.row == 2 or cell.row == 5:  # Special rows (Agent Total and Total Invoice)
                    cell.font = Font(bold=True)
                    cell.fill = row_fill

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "Airtel Invoice"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = header_fill

        wb.save(file_path)

        return file_path
    except pd.errors.EmptyDataError:
        st.error("No columns to parse from file or the file is empty.")
        return None
    except Exception as e:
        st.error(f"An error occurred while processing the file: {e}")
        return None

def main():
    st.title("Airtel Invoice Generator")
    
    uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

    if uploaded_file is not None:
        if st.button("Generate Output"):
            with open("uploaded_file.csv", "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            output_file_path = process_data("uploaded_file.csv")
            if output_file_path:
                st.success(f"Summary table saved to: {output_file_path}")
                with open(output_file_path, "rb") as f:
                    st.download_button(
                        label="Download Invoice",
                        data=f,
                        file_name=output_file_path,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

if __name__ == "__main__":
    main()
